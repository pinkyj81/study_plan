# app.py — DB 연결 Flask 학습캘린더
from flask import Flask, render_template, jsonify, request, session, redirect, url_for
from datetime import date, datetime, timedelta
from calendar import monthrange
from db_config import db, init_db
from sqlalchemy import text
from functools import wraps
from io import TextIOWrapper
import csv
import re

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'  # 세션용 비밀키
init_db(app)

# 파스텔 색상 팔레트
PLAN_COLORS = [
    '#FFB3BA',  # pastel pink
    '#FFDFBA',  # pastel orange
    '#FFFFBA',  # pastel yellow
    '#BAFFC9',  # pastel green
    '#BAE1FF',  # pastel blue
    '#D4BEEE',  # deeper purple
    '#FFDFD3',  # pastel peach
]

# 로그인 체크 데코레이터
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def month_days(year, month):
    """앞쪽 공백 포함 달력 데이터 생성"""
    first_weekday, last_day = monthrange(year, month)
    start_padding = (first_weekday + 1) % 7
    days = [None] * start_padding
    days += [date(year, month, d) for d in range(1, last_day + 1)]
    return days

# DB에서 학습계획 가져오기
def get_plans_from_db():
    """데이터베이스에서 학습계획 조회"""
    # 한 번의 쿼리로 모든 데이터 조회 (JOIN 사용)
    query = text("""
        SELECT 
            p.plan_id,
            p.title,
            p.subject,
            p.created_at,
            t.task_id,
            t.plan_date,
            t.task_title,
            t.order_no,
            ISNULL(
                (SELECT TOP 1 status 
                 FROM dbo.study_plan_log 
                 WHERE task_id = t.task_id 
                 ORDER BY updated_at DESC), 
                'planned'
            ) as status
        FROM dbo.study_plan p
        LEFT JOIN dbo.study_plan_task t ON p.plan_id = t.plan_id
        ORDER BY p.created_at DESC, t.plan_date, t.order_no
    """)
    
    result = db.session.execute(query).fetchall()
    
    # 결과를 계획별로 그룹화
    plans_dict = {}
    
    for row in result:
        plan_id = row.plan_id
        
        # 새로운 계획인 경우
        if plan_id not in plans_dict:
            # 색상 할당 (plan_id 기반으로 순환)
            color_idx = (plan_id - 1) % len(PLAN_COLORS)
            plans_dict[plan_id] = {
                "plan_id": plan_id,
                "title": row.title,
                "subject": row.subject,
                "created_at": row.created_at.strftime("%Y-%m-%d") if row.created_at else "",
                "color": PLAN_COLORS[color_idx],
                "days": [],  # 호환성 유지
                "daily_plans": []
            }
        
        # 일일 작업이 있는 경우
        if row.task_id:
            plans_dict[plan_id]["daily_plans"].append({
                "task_id": row.task_id,
                "date": row.plan_date.strftime("%Y-%m-%d"),
                "order": row.order_no,
                "description": row.task_title,
                "status": row.status
            })
    
    return list(plans_dict.values())

def generate_fake_calendar(year, plan_id=None):
    data = {}
    today = datetime.now().date()
    
    # 세션에서 user_id 가져오기
    user_id = session.get('user_id', 1)
    
    # DB에서 학습계획 가져오기
    plans = get_plans_from_db()
    
    # 모든 일일 계획을 날짜별로 매핑 (색상 포함)
    daily_plan_map = {}
    for plan in plans:
        # 특정 계획 필터링: plan_id가 지정된 경우 해당 계획만 반영
        if plan_id and plan.get("plan_id") != plan_id:
            continue
        plan_color = plan.get("color", "#E5E7EB")  # 기본 회색
        plan_id_num = plan.get("plan_id")
        for dp in plan.get("daily_plans", []):
            date_str = dp.get("date")
            status = dp.get("status", "planned")  # planned, done, partial, missed
            if date_str:
                daily_plan_map[date_str] = {
                    "status": status,
                    "color": plan_color,
                    "plan_id": plan_id_num
                }
    
    for m in range(1, 13):
        month_list = []
        for d in month_days(year, m):
            if d is None:
                month_list.append(None)
            else:
                date_str = d.strftime("%Y-%m-%d")
                
                # 일일 계획에서 상태와 색상 확인
                color = None
                plan_id_val = None
                if date_str in daily_plan_map:
                    status = daily_plan_map[date_str]["status"]
                    color = daily_plan_map[date_str]["color"]
                    plan_id_val = daily_plan_map[date_str]["plan_id"]
                elif d < today:
                    status = "none"  # 계획 없음 (과거)
                else:
                    status = "none"
                
                # 오늘 표시
                is_today = (d == today)
                
                month_list.append({
                    "date": d,
                    "status": status,
                    "is_today": is_today,
                    "day_id": f"{m:02d}{d.day:02d}",
                    "color": color,
                    "plan_id": plan_id_val
                })
        data[m] = month_list
    return data

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        data = request.get_json(force=True) if request.is_json else request.form
        user_name = data.get("username")
        
        try:
            # 사용자 확인
            query = text("SELECT user_id, user_name FROM dbo.study_plan_user WHERE user_name = :username")
            result = db.session.execute(query, {"username": user_name}).fetchone()
            
            if result:
                # 로그인 성공
                session['user_id'] = result.user_id
                session['user_name'] = result.user_name
                
                if request.is_json:
                    return jsonify({"ok": True, "message": "로그인 성공!"})
                return redirect(url_for('index'))
            else:
                # 사용자가 없으면 자동 생성
                insert_query = text("""
                    INSERT INTO dbo.study_plan_user (user_name, created_at)
                    VALUES (:username, SYSDATETIMEOFFSET())
                """)
                db.session.execute(insert_query, {"username": user_name})
                db.session.commit()
                
                # 새로 생성된 사용자 조회
                result = db.session.execute(query, {"username": user_name}).fetchone()
                session['user_id'] = result.user_id
                session['user_name'] = result.user_name
                
                if request.is_json:
                    return jsonify({"ok": True, "message": "계정이 생성되었습니다!"})
                return redirect(url_for('index'))
                
        except Exception as e:
            db.session.rollback()
            if request.is_json:
                return jsonify({"ok": False, "error": str(e)}), 500
            return render_template("login.html", error=str(e))
    
    # GET 요청 - 로그인 페이지 표시
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route("/")
@app.route("/<int:year>")
@login_required
def index(year=2025):
    # DB에서 학습계획 가져오기
    plans = get_plans_from_db()
    # 선택된 계획 필터 (쿼리 파라미터)
    active_plan_id = request.args.get('plan_id', type=int)
    active_plan = None
    if active_plan_id:
        active_plan = next((p for p in plans if p.get('plan_id') == active_plan_id), None)
    
    # 년도별 달력 데이터 생성 (선택된 계획 기준으로 색상 상태 반영)
    calendar_data = generate_fake_calendar(year, plan_id=active_plan_id)
    
    # 년도별 요약 데이터 (예시)
    days_in_year = 366 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 365
    summary = {
        "total": days_in_year,
        "passed": days_in_year // 2 if year < 2025 else 294,
        "remain": days_in_year - (days_in_year // 2 if year < 2025 else 294),
        "planned": 48
    }
    
    # 통계 계산 (선택된 계획이 있으면 해당 계획 기준)
    base_list = [active_plan] if active_plan else plans
    total_plans = len(base_list)
    total_assigned = sum(len(p["daily_plans"]) for p in base_list)
    completed = sum(1 for p in base_list for dp in p["daily_plans"] if dp.get("status") == "done")
    completion_rate = int((completed / total_assigned * 100)) if total_assigned > 0 else 0
    
    stats = {
        "total_plans": total_plans,
        "total_assigned": total_assigned,
        "completed": completed,
        "completion_rate": completion_rate
    }
    
    return render_template(
        "index.html",
        year=year,
        available_years=[2024, 2025, 2026, 2027],
        plans=plans,
        active_plan=active_plan or (plans[0] if plans else None),
        calendar_data=calendar_data,
        summary=summary,
        stats=stats,
        today_date=date.today()
    )

@app.route("/manage_plan")
@login_required
def manage_plan():
    # DB에서 학습계획 가져오기
    plans = get_plans_from_db()
    
    # 통계 계산
    total_plans = len(plans)
    total_assigned = sum(len(p["daily_plans"]) for p in plans)
    completed = sum(1 for p in plans for dp in p["daily_plans"] if dp.get("status") == "done")
    completion_rate = int((completed / total_assigned * 100)) if total_assigned > 0 else 0
    
    stats = {
        "total_plans": total_plans,
        "total_assigned": total_assigned,
        "completed": completed,
        "completion_rate": completion_rate
    }
    
    # JSON으로 변환하기 위해 import 추가
    import json
    
    return render_template(
        "manage_plan.html",
        plans=plans,
        plans_json=json.dumps(plans),
        stats=stats
    )

@app.route("/day/<day_id>")
def day_detail(day_id):
    # day_id 형식: "MMDD" (예: "1022" = 10월 22일)
    year = request.args.get('year', 2025, type=int)
    try:
        month = int(day_id[:2])
        day = int(day_id[2:])
        
        # 날짜 유효성 검사
        date_obj = date(year, month, day)
        date_str = date_obj.strftime("%Y-%m-%d")
        
        # DB에서 해당 날짜의 학습 작업 조회 (선택된 계획만 필터 가능)
        plan_id_param = request.args.get('plan_id', type=int)
        query = text("""
            SELECT 
                p.title as plan_title,
                p.subject,
                t.task_id,
                t.task_title,
                t.link_url,
                ISNULL(
                    (SELECT TOP 1 status 
                     FROM dbo.study_plan_log 
                     WHERE task_id = t.task_id 
                     ORDER BY updated_at DESC), 
                    'planned'
                ) as status,
                ISNULL(
                    (SELECT TOP 1 actual_minutes 
                     FROM dbo.study_plan_log 
                     WHERE task_id = t.task_id 
                     ORDER BY updated_at DESC), 
                    0
                ) as minutes,
                ISNULL(
                    (SELECT TOP 1 memo 
                     FROM dbo.study_plan_log 
                     WHERE task_id = t.task_id 
                     ORDER BY updated_at DESC), 
                    ''
                ) as memo
            FROM dbo.study_plan_task t
            JOIN dbo.study_plan p ON t.plan_id = p.plan_id
            WHERE t.plan_date = :date
              AND (:plan_id IS NULL OR p.plan_id = :plan_id)
            ORDER BY t.order_no
        """)
        
        result = db.session.execute(query, {"date": date_str, "plan_id": plan_id_param}).fetchall()
        
        # 해당 날짜에 작업이 있는 경우
        if result:
            tasks = []
            for row in result:
                tasks.append({
                    "task_id": row.task_id,
                    "plan_title": row.plan_title,
                    "subject": row.subject,
                    "task_title": row.task_title,
                    "link_url": row.link_url,
                    "status": row.status
                })
            return jsonify({
                "ok": True,
                "date": date_str,
                "tasks": tasks
            })
        else:
            # 작업이 없는 경우
            return jsonify({
                "ok": True,
                "date": date_str,
                "tasks": []
            })
            
    except (ValueError, IndexError) as e:
        return jsonify({
            "ok": False,
            "error": "잘못된 날짜 형식입니다"
        })
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/day/update", methods=["POST"])
def day_update():
    data = request.get_json(force=True)
    
    try:
        task_id = data.get("task_id")
        is_completed = data.get("completed", False)
        
        if not task_id:
            return jsonify({
                "ok": False,
                "error": "task_id가 필요합니다."
            }), 400
        
        # 완료 체크박스가 체크되면 done, 아니면 planned
        status = "done" if is_completed else "planned"
        
        # user_id = 1 확인 및 생성
        user_check = text("SELECT user_id FROM dbo.study_plan_user WHERE user_id = 1")
        user_exists = db.session.execute(user_check).fetchone()
        
        if not user_exists:
            # 기본 사용자 생성 (user_name 포함)
            create_user = text("""
                SET IDENTITY_INSERT dbo.study_plan_user ON;
                INSERT INTO dbo.study_plan_user (user_id, user_name, created_at) 
                VALUES (1, 'default_user', SYSDATETIMEOFFSET());
                SET IDENTITY_INSERT dbo.study_plan_user OFF;
            """)
            db.session.execute(create_user)
            db.session.commit()
        
        # 기존 로그가 있는지 확인
        check_query = text("""
            SELECT log_id 
            FROM dbo.study_plan_log 
            WHERE task_id = :task_id
        """)
        existing_log = db.session.execute(check_query, {"task_id": task_id}).fetchone()
        
        if existing_log:
            # 기존 로그 업데이트
            update_query = text("""
                UPDATE dbo.study_plan_log
                SET status = :status,
                    updated_at = SYSDATETIMEOFFSET()
                WHERE task_id = :task_id
            """)
            db.session.execute(update_query, {
                "task_id": task_id,
                "status": status
            })
        else:
            # 새 로그 삽입
            insert_query = text("""
                INSERT INTO dbo.study_plan_log 
                (task_id, user_id, status, updated_at)
                VALUES (:task_id, 1, :status, SYSDATETIMEOFFSET())
            """)
            db.session.execute(insert_query, {
                "task_id": task_id,
                "status": status
            })
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": "학습 실적이 저장되었습니다!"
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/plan/create", methods=["POST"])
def create_plan():
    data = request.get_json(force=True)
    
    try:
        # DB에 새 계획 추가
        insert_query = text("""
            INSERT INTO dbo.study_plan (title, subject, created_at)
            VALUES (:title, :subject, SYSDATETIMEOFFSET())
        """)
        
        db.session.execute(insert_query, {
            "title": data["title"],
            "subject": data["subject"]
        })
        db.session.commit()
        
        # 새로 생성된 plan_id 조회
        id_query = text("SELECT MAX(plan_id) as new_id FROM dbo.study_plan")
        result = db.session.execute(id_query).fetchone()
        new_plan_id = result.new_id
        
        return jsonify({
            "ok": True,
            "message": "계획이 성공적으로 추가되었습니다!",
            "plan_id": new_plan_id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/plan/create_from_template", methods=["POST"])
def create_plan_from_template():
    """템플릿에서 새 계획 생성: 템플릿 항목들을 날짜 범위에 맞춰 일일 계획으로 변환"""
    data = request.get_json(force=True)
    
    try:
        source_template_id = data.get("source_template_id")
        title = data.get("title")
        subject = data.get("subject")
        start_date_str = data.get("start_date")
        end_date_str = data.get("end_date")
        
        if not all([source_template_id, title, subject, start_date_str, end_date_str]):
            return jsonify({"ok": False, "error": "필수 항목이 누락되었습니다."}), 400
        
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        
        if start_date > end_date:
            return jsonify({"ok": False, "error": "시작일이 종료일보다 늦습니다."}), 400
        
        # 1. 새 계획 생성
        insert_plan = text("""
            INSERT INTO dbo.study_plan (title, subject, created_at)
            VALUES (:title, :subject, SYSDATETIMEOFFSET())
        """)
        db.session.execute(insert_plan, {"title": title, "subject": subject})
        db.session.commit()
        
        # 새 plan_id 가져오기
        id_query = text("SELECT MAX(plan_id) as new_id FROM dbo.study_plan")
        new_plan_id = db.session.execute(id_query).fetchone().new_id
        
        # 2. 선택한 템플릿의 항목들 가져오기 (독립 템플릿 스키마)
        template_query = text("""
            SELECT order_no, title, link_url
            FROM dbo.study_task_template
            WHERE template_id = :template_id
            ORDER BY order_no
        """)
        templates = db.session.execute(template_query, {"template_id": source_template_id}).fetchall()
        
        if not templates:
            return jsonify({"ok": False, "error": "선택한 템플릿에 항목이 없습니다."}), 400
        
        # 3. 날짜 범위 계산
        total_days = (end_date - start_date).days + 1
        template_count = len(templates)
        
        # 템플릿을 날짜에 균등 분배
        current_date = start_date
        task_insert = text("""
            INSERT INTO dbo.study_plan_task (plan_id, plan_date, task_title, order_no, link_url, created_at)
            VALUES (:plan_id, :plan_date, :task_title, :order_no, :link_url, SYSDATETIMEOFFSET())
        """)
        
        created_count = 0
        for idx, template in enumerate(templates):
            if current_date > end_date:
                break
            
            db.session.execute(task_insert, {
                "plan_id": new_plan_id,
                "plan_date": current_date,
                "task_title": template.title,
                "order_no": idx + 1,
                "link_url": template.link_url
            })
            created_count += 1
            
            # 다음 날짜로 이동
            current_date = current_date + timedelta(days=1)
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": f"'{title}' 계획이 생성되었습니다!",
            "plan_id": new_plan_id,
            "count": created_count
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/plan/<int:plan_id>/daily", methods=["POST"])
def save_daily_plans(plan_id):
    data = request.get_json(force=True)
    daily_plans = data.get("daily_plans", [])
    
    try:
        # 기존 로그 삭제 (외래키 관계)
        log_delete_query = text("""
            DELETE FROM dbo.study_plan_log
            WHERE task_id IN (
                SELECT task_id FROM dbo.study_plan_task WHERE plan_id = :plan_id
            )
        """)
        db.session.execute(log_delete_query, {"plan_id": plan_id})
        
        # 기존 일일 작업 삭제
        delete_query = text("DELETE FROM dbo.study_plan_task WHERE plan_id = :plan_id")
        db.session.execute(delete_query, {"plan_id": plan_id})
        
        # 새 일일 작업 추가 및 task_id 수집
        for idx, plan in enumerate(daily_plans):
            insert_query = text("""
                INSERT INTO dbo.study_plan_task 
                (plan_id, plan_date, task_title, link_url, order_no, created_at)
                OUTPUT INSERTED.task_id
                VALUES (:plan_id, :plan_date, :task_title, :link_url, :order_no, SYSDATETIMEOFFSET())
            """)
            
            result = db.session.execute(insert_query, {
                "plan_id": plan_id,
                "plan_date": plan.get("date"),
                "task_title": plan.get("description", ""),
                "link_url": plan.get("link_url"),
                "order_no": plan.get("order", idx + 1)
            })
            
            # SCOPE_IDENTITY()를 통해 방금 삽입한 task_id 가져오기
            task_id = result.fetchone()[0]
            
            # 로그 저장 (상태가 있는 경우)
            if plan.get("status") and plan.get("status") != "planned":
                log_insert_query = text("""
                    INSERT INTO dbo.study_plan_log 
                    (task_id, user_id, status, actual_minutes, memo, updated_at)
                    VALUES (:task_id, 1, :status, 0, '', SYSDATETIMEOFFSET())
                """)
                db.session.execute(log_insert_query, {
                    "task_id": task_id,
                    "status": plan.get("status")
                })
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": "일일 계획이 저장되었습니다!",
            "plan_id": plan_id,
            "count": len(daily_plans)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/plan/<int:plan_id>/daily", methods=["GET"])
def get_daily_plans(plan_id):
    try:
        # DB에서 일일 작업 조회
        query = text("""
            SELECT 
                t.task_id,
                t.plan_date,
                t.task_title,
                t.link_url,
                t.order_no,
                ISNULL(l.status, 'planned') as status
            FROM dbo.study_plan_task t
            LEFT JOIN (
                SELECT task_id, status,
                       ROW_NUMBER() OVER (PARTITION BY task_id ORDER BY updated_at DESC) as rn
                FROM dbo.study_plan_log
            ) l ON t.task_id = l.task_id AND l.rn = 1
            WHERE t.plan_id = :plan_id
            ORDER BY t.plan_date, t.order_no
        """)
        
        result = db.session.execute(query, {"plan_id": plan_id})
        daily_plans = []
        
        for row in result:
            daily_plans.append({
                "task_id": row.task_id,
                "date": row.plan_date.strftime("%Y-%m-%d"),
                "description": row.task_title,
                "link_url": row.link_url,
                "order": row.order_no,
                "status": row.status
            })
        
        return jsonify({
            "ok": True,
            "daily_plans": daily_plans
        })
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/plan/<int:plan_id>/update", methods=["POST"])
def update_plan(plan_id):
    data = request.get_json(force=True)
    
    try:
        # DB에서 계획 업데이트
        update_query = text("""
            UPDATE dbo.study_plan
            SET title = :title,
                subject = :subject
            WHERE plan_id = :plan_id
        """)
        
        db.session.execute(update_query, {
            "plan_id": plan_id,
            "title": data.get("title"),
            "subject": data.get("subject")
        })
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": "계획이 성공적으로 수정되었습니다!"
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@app.route("/plan/<int:plan_id>/delete", methods=["POST"])
def delete_plan(plan_id):
    try:
        # 계획 제목 조회 (메시지용)
        title_query = text("SELECT title FROM dbo.study_plan WHERE plan_id = :plan_id")
        title_result = db.session.execute(title_query, {"plan_id": plan_id}).fetchone()
        
        if not title_result:
            return jsonify({"ok": False, "error": "계획을 찾을 수 없습니다"}), 404
        
        plan_title = title_result.title
        
        # 관련 로그 삭제 (외래키 관계)
        log_delete_query = text("""
            DELETE FROM dbo.study_plan_log
            WHERE task_id IN (
                SELECT task_id FROM dbo.study_plan_task WHERE plan_id = :plan_id
            )
        """)
        db.session.execute(log_delete_query, {"plan_id": plan_id})
        
        # 일일 작업 삭제
        task_delete_query = text("DELETE FROM dbo.study_plan_task WHERE plan_id = :plan_id")
        db.session.execute(task_delete_query, {"plan_id": plan_id})
        
        # 계획 삭제
        plan_delete_query = text("DELETE FROM dbo.study_plan WHERE plan_id = :plan_id")
        db.session.execute(plan_delete_query, {"plan_id": plan_id})
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "message": f'"{plan_title}" 계획이 삭제되었습니다!'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

# ======================= 템플릿 관리 API =======================

@app.route("/templates/create", methods=["POST"])
def create_template():
    """새 템플릿 생성: 템플릿 + 템플릿 항목들"""
    try:
        data = request.get_json(force=True)
        template_title = data.get("template_title", "").strip()
        subject = data.get("subject")
        description = data.get("description")
        items = data.get("items", [])
        
        if not template_title:
            return jsonify({"ok": False, "error": "템플릿 제목이 필요합니다."}), 400
        
        if not items or not isinstance(items, list):
            return jsonify({"ok": False, "error": "템플릿 항목이 필요합니다."}), 400
        
        # 1. 템플릿 생성
        template_insert = text("""
            INSERT INTO dbo.study_template (template_title, subject, description, created_at)
            VALUES (:title, :subject, :description, SYSDATETIMEOFFSET())
        """)
        db.session.execute(template_insert, {
            "title": template_title,
            "subject": subject,
            "description": description
        })
        db.session.commit()
        
        # 2. 생성된 template_id 가져오기
        id_query = text("SELECT MAX(template_id) as new_id FROM dbo.study_template")
        new_template_id = db.session.execute(id_query).fetchone().new_id
        
        # 3. 템플릿 항목들 추가
        item_insert = text("""
            INSERT INTO dbo.study_task_template (template_id, order_no, title, link_url, created_at)
            VALUES (:template_id, :order_no, :title, :link_url, SYSDATETIMEOFFSET())
        """)
        
        for item in items:
            db.session.execute(item_insert, {
                "template_id": new_template_id,
                "order_no": item.get("order_no", 1),
                "title": item.get("title", "").strip(),
                "link_url": item.get("link_url")
            })
        
        db.session.commit()
        
        return jsonify({
            "ok": True,
            "template_id": new_template_id,
            "item_count": len(items)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/templates", methods=["GET"])
def get_all_templates():
    """모든 템플릿 목록 조회 (독립 템플릿)"""
    try:
        query = text(
            """
            SELECT template_id, template_title, subject, description, created_at
            FROM dbo.study_template
            ORDER BY created_at DESC
            """
        )
        rows = db.session.execute(query).fetchall()
        templates = [
            {
                "template_id": r.template_id,
                "template_title": r.template_title,
                "subject": r.subject,
                "description": r.description,
                "created_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else None,
            }
            for r in rows
        ]
        return jsonify({"ok": True, "templates": templates})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/templates/<int:template_id>/items", methods=["GET"])
def get_template_items(template_id: int):
    """특정 템플릿의 항목들 조회"""
    try:
        query = text(
            """
            SELECT order_no, title, link_url
            FROM dbo.study_task_template
            WHERE template_id = :template_id
            ORDER BY order_no
            """
        )
        rows = db.session.execute(query, {"template_id": template_id}).fetchall()
        items = [
            {
                "order_no": r.order_no,
                "title": r.title,
                "link_url": r.link_url,
            }
            for r in rows
        ]
        return jsonify({"ok": True, "items": items})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/templates/<int:template_id>/items/add", methods=["POST"])
def add_template_items(template_id: int):
    """템플릿에 항목들 추가"""
    try:
        data = request.get_json(force=True)
        items = data.get("items", [])
        
        if not items:
            return jsonify({"ok": False, "error": "추가할 항목이 없습니다."}), 400
        
        insert_q = text(
            """
            INSERT INTO dbo.study_task_template (template_id, order_no, title, link_url, created_at)
            VALUES (:template_id, :order_no, :title, :link_url, SYSDATETIMEOFFSET())
            """
        )
        
        count = 0
        for item in items:
            title = (item.get("title") or "").strip()
            if not title:
                continue
            
            db.session.execute(insert_q, {
                "template_id": template_id,
                "order_no": item.get("order_no", count + 1),
                "title": title,
                "link_url": item.get("link_url")
            })
            count += 1
        
        db.session.commit()
        return jsonify({"ok": True, "count": count})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/templates/<int:template_id>/items/bulk_update", methods=["POST"])
def bulk_update_template_items(template_id: int):
    """템플릿 항목들 일괄 수정"""
    try:
        data = request.get_json(force=True)
        items = data.get("items", [])
        
        if not items:
            return jsonify({"ok": False, "error": "수정할 항목이 없습니다."}), 400
        
        # 먼저 현재 템플릿의 모든 항목 삭제 후 새로 추가
        # (중복 방지)
        del_q = text("DELETE FROM dbo.study_task_template WHERE template_id = :template_id")
        db.session.execute(del_q, {"template_id": template_id})
        
        # 새 항목 추가
        ins_q = text(
            """
            INSERT INTO dbo.study_task_template (template_id, order_no, title, link_url, created_at)
            VALUES (:template_id, :order_no, :title, :link_url, SYSDATETIMEOFFSET())
            """
        )
        
        count = 0
        for item in items:
            title = (item.get("title") or "").strip()
            link = item.get("link_url")
            order = item.get("order_no")
            
            if not title:
                continue
            
            try:
                order = int(order) if order is not None else count + 1
            except:
                order = count + 1
            
            db.session.execute(ins_q, {
                "template_id": template_id,
                "title": title,
                "link_url": link,
                "order_no": order
            })
            count += 1
        
        db.session.commit()
        return jsonify({"ok": True, "count": count})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/templates/<int:template_id>/items/<int:item_id>/delete", methods=["POST"])
def delete_template_item(template_id: int, item_id: int):
    """템플릿 항목 삭제"""
    try:
        del_q = text(
            """
            DELETE FROM dbo.study_task_template
            WHERE template_id = :template_id
            """
        )
        db.session.execute(del_q, {"template_id": template_id})
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


def _parse_paste_text(paste_text: str):
    """붙여넣기 텍스트를 행(row) 리스트로 파싱
    지원 포맷: title, link_url (콤마 또는 탭 구분, memo 제외)
    """
    rows = []
    if not paste_text:
        return rows
    lines = [ln.strip() for ln in paste_text.splitlines() if ln.strip()]
    for idx, ln in enumerate(lines, start=1):
        # 탭 또는 콤마로 분리: 업로드/붙여넣기에서는 memo를 사용하지 않음
        parts = re.split(r"\t|,", ln)
        title = parts[0].strip() if len(parts) > 0 else ""
        link = parts[1].strip() if len(parts) > 1 else None
        if title:
            rows.append({"order_no": idx, "title": title, "link_url": link})
    return rows


def _parse_csv_file(file_stream):
    """CSV 파일 스트림을 읽어 rows 반환. 헤더(optional): title,link_url"""
    reader = csv.DictReader(TextIOWrapper(file_stream, encoding="utf-8"))
    rows = []
    for i, row in enumerate(reader, start=1):
        title = (row.get("title") or row.get("Title") or "").strip()
        link = (row.get("link_url") or row.get("link") or row.get("Link"))
        if title:
            rows.append({"order_no": i, "title": title, "link_url": link})
    return rows


@app.route("/plan/<int:plan_id>/templates/upload", methods=["POST"])
def upload_templates(plan_id: int):
    """템플릿 업로드: CSV/Excel 파일 또는 붙여넣기 텍스트 지원

    요청 형태:
    - multipart/form-data: file(필수), csv 또는 xlsx 확장자
    - application/json: { paste_text: "제목\t링크\n..." } 또는 { rows: [{title, link_url, order_no}, ...] }
    """
    try:
        # 현재 마지막 order_no 조회
        max_order_q = text(
            "SELECT ISNULL(MAX(order_no), 0) AS max_order FROM dbo.study_task_template WHERE plan_id = :plan_id"
        )
        max_order = db.session.execute(max_order_q, {"plan_id": plan_id}).fetchone().max_order or 0

        rows_to_insert = []

        if request.content_type and request.content_type.startswith("multipart/form-data"):
            file = request.files.get("file")
            if not file or not file.filename:
                return jsonify({"ok": False, "error": "파일을 선택하세요."}), 400
            filename = file.filename.lower()
            if filename.endswith(".csv"):
                rows_to_insert = _parse_csv_file(file.stream)
            elif filename.endswith(".xlsx"):
                try:
                    import openpyxl  # optional dependency

                    wb = openpyxl.load_workbook(file)
                    ws = wb.active
                    # 첫 행이 헤더라고 가정: title, link_url (memo 제외)
                    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
                    col_title = headers.index("title") if "title" in headers else 0
                    col_link = headers.index("link_url") if "link_url" in headers else None
                    order = 1
                    for row in ws.iter_rows(min_row=2):
                        title = str(row[col_title].value).strip() if row[col_title].value else ""
                        link = str(row[col_link].value).strip() if (col_link is not None and row[col_link].value) else None
                        if title:
                            rows_to_insert.append({"order_no": order, "title": title, "link_url": link})
                            order += 1
                except ImportError:
                    return jsonify({"ok": False, "error": "Excel(.xlsx) 업로드는 openpyxl 설치가 필요합니다."}), 400
            else:
                return jsonify({"ok": False, "error": "지원되지 않는 파일 형식입니다. .csv 또는 .xlsx 사용"}), 400

        elif request.is_json:
            data = request.get_json(force=True)
            if isinstance(data.get("rows"), list):
                rows_to_insert = [
                    {
                        "order_no": r.get("order_no"),
                        "title": r.get("title", "").strip(),
                        "link_url": r.get("link_url"),
                    }
                    for r in data.get("rows", [])
                    if r.get("title")
                ]
            else:
                paste_text = data.get("paste_text", "")
                rows_to_insert = _parse_paste_text(paste_text)
        else:
            return jsonify({"ok": False, "error": "잘못된 요청 형식입니다."}), 400

        if not rows_to_insert:
            return jsonify({"ok": False, "error": "추가할 행이 없습니다."}), 400

        # 삽입 실행
        insert_q = text(
            """
            INSERT INTO dbo.study_task_template (plan_id, order_no, title, link_url, created_at)
            VALUES (:plan_id, :order_no, :title, :link_url, SYSDATETIMEOFFSET())
            """
        )

        order_cursor = max_order
        for r in rows_to_insert:
            order_cursor += 1 if not r.get("order_no") else 0
            order_no = r.get("order_no") or order_cursor
            db.session.execute(
                insert_q,
                {
                    "plan_id": plan_id,
                    "order_no": order_no,
                    "title": r.get("title"),
                    "link_url": r.get("link_url"),
                },
            )

        db.session.commit()
        return jsonify({"ok": True, "count": len(rows_to_insert)})

    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# 단일 템플릿 수정
@app.route("/plan/<int:plan_id>/templates/<int:template_id>/update", methods=["POST"])
def update_template(plan_id: int, template_id: int):
    """템플릿 한 건 수정: title, link_url, order_no 지원"""
    try:
        data = request.get_json(force=True) if request.is_json else request.form

        # 기존 레코드 확인 (plan_id 일치 검증)
        sel_q = text(
            """
            SELECT template_id, plan_id, order_no, title, link_url
            FROM dbo.study_task_template
            WHERE template_id = :template_id AND plan_id = :plan_id
            """
        )
        existing = db.session.execute(sel_q, {"template_id": template_id, "plan_id": plan_id}).fetchone()
        if not existing:
            return jsonify({"ok": False, "error": "템플릿을 찾을 수 없습니다."}), 404

        new_title = (data.get("title") or existing.title).strip()
        new_link = data.get("link_url") if data.get("link_url") is not None else existing.link_url
        new_order = data.get("order_no")
        try:
            new_order = int(new_order) if new_order is not None else existing.order_no
        except Exception:
            new_order = existing.order_no

        upd_q = text(
            """
            UPDATE dbo.study_task_template
            SET title = :title, link_url = :link_url, order_no = :order_no
            WHERE template_id = :template_id AND plan_id = :plan_id
            """
        )
        db.session.execute(
            upd_q,
            {
                "title": new_title,
                "link_url": new_link,
                "order_no": new_order,
                "template_id": template_id,
                "plan_id": plan_id,
            },
        )
        db.session.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# 단일 템플릿 삭제
@app.route("/plan/<int:plan_id>/templates/<int:template_id>/delete", methods=["POST"])
def delete_template(plan_id: int, template_id: int):
    """템플릿 한 건 삭제"""
    try:
        del_q = text(
            """
            DELETE FROM dbo.study_task_template
            WHERE template_id = :template_id AND plan_id = :plan_id
            """
        )
        res = db.session.execute(del_q, {"template_id": template_id, "plan_id": plan_id})
        db.session.commit()
        # res.rowcount 가 0이면 없던 것
        if getattr(res, "rowcount", None) == 0:
            return jsonify({"ok": False, "error": "삭제할 템플릿이 없습니다."}), 404
        return jsonify({"ok": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# 템플릿 일괄 수정
@app.route("/plan/<int:plan_id>/templates/bulk_update", methods=["POST"])
def bulk_update_templates(plan_id: int):
    """템플릿 다건 수정: [{template_id, title, link_url, order_no}, ...]"""
    try:
        if not request.is_json:
            return jsonify({"ok": False, "error": "JSON 요청이 필요합니다."}), 400
        data = request.get_json(force=True)
        templates = data.get("templates", [])
        if not isinstance(templates, list) or not templates:
            return jsonify({"ok": False, "error": "수정할 항목이 없습니다."}), 400

        upd_q = text(
            """
            UPDATE dbo.study_task_template
            SET title = :title, link_url = :link_url, order_no = :order_no
            WHERE template_id = :template_id AND plan_id = :plan_id
            """
        )

        count = 0
        for r in templates:
            tid = r.get("template_id")
            title = (r.get("title") or "").strip()
            link = r.get("link_url")
            order = r.get("order_no")
            try:
                order = int(order) if order is not None else None
            except Exception:
                order = None
            if not tid or not title or order is None:
                # 필수값 누락 시 건너뜀
                continue
            db.session.execute(
                upd_q,
                {
                    "title": title,
                    "link_url": link,
                    "order_no": order,
                    "template_id": tid,
                    "plan_id": plan_id,
                },
            )
            count += 1

        db.session.commit()
        return jsonify({"ok": True, "count": count})
    except Exception as e:
        db.session.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# 템플릿 관리 전용 페이지
@app.route("/templates/manage")
@login_required
def template_manage_page():
    plans = get_plans_from_db()
    active_plan_id = request.args.get("plan_id", type=int)
    if not active_plan_id and plans:
        active_plan_id = plans[0]["plan_id"]
    return render_template("template_manage.html", plans=plans, active_plan_id=active_plan_id)
if __name__ == "__main__":
    app.run(debug=True)
